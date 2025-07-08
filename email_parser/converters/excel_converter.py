"""
Excel converter module for converting Excel workbooks to CSV files.
"""

import logging
import os
from pathlib import Path

# from typing import Dict, List, Optional, Tuple, Any, Callable
from typing import Any, Callable, Dict, List, Optional

import pandas as pd  # type: ignore
from openpyxl import load_workbook  # type: ignore

from email_parser.exceptions.parsing_exceptions import ExcelConversionError

# from email_parser.utils.file_utils import ensure_directory, generate_unique_filename
# from email_parser.utils.file_utils import ensure_directory
from email_parser.utils.file_utils import ensure_directory, generate_unique_filename

logger = logging.getLogger(__name__)


class ExcelConverter:
    """
    Converts Excel workbook attachments to CSV files.

    This class handles the detection and conversion of Excel files (.xlsx, .xls)
    to CSV format, with options for user interaction.
    """

    def __init__(self, output_dir: str = "output/converted_excel"):
        """
        Initialize the Excel converter.

        Args:
            output_dir: Directory for saving converted CSV files
        """
        self.output_dir = output_dir
        ensure_directory(output_dir)

    def is_excel_file(self, filename: str, content_type: Optional[str] = None) -> bool:
        """
        Determine if a file is an Excel workbook.

        Args:
            filename: Name of the file
            content_type: MIME content type if available

        Returns:
            True if the file appears to be an Excel workbook, False otherwise
        """
        if filename:
            _, ext = os.path.splitext(filename)
            if ext.lower() in (".xlsx", ".xls", ".xlsm", ".xlsb"):
                return True

        if content_type:
            excel_mime_types = [
                "application/vnd.ms-excel",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "application/vnd.ms-excel.sheet.macroEnabled.12",
                "application/vnd.ms-excel.sheet.binary.macroEnabled.12",
                "application/octet-stream",  # Added generic MIME type
            ]

            if content_type in excel_mime_types:
                # For octet-stream, we should check the filename too
                if content_type == "application/octet-stream":
                    if filename:
                        _, ext = os.path.splitext(filename)
                        return ext.lower() in (".xlsx", ".xls", ".xlsm", ".xlsb")
                    return False
                return True

        return False

    def convert_excel_to_csv(
        self,
        excel_path: str,
        original_filename: str,
        secure_filename: str,
        email_id: str,
        prompt_callback: Optional[Callable[[str, List[str]], List[str]]] = None,
    ) -> List[Dict[str, Any]]:
        """
        Convert an Excel workbook to CSV files.

        Args:
            excel_path: Path to the Excel file
            original_filename: Original filename of the Excel file
            secure_filename: Secure filename of the Excel file
            email_id: Unique identifier for the email
            prompt_callback: Optional callback function for user prompting

        Returns:
            List of dictionaries with information about converted CSV files

        Raises:
            ExcelConversionError: If conversion fails
        """
        conversion_results = []

        try:
            # Get list of sheet names
            sheet_names = self._get_sheet_names(excel_path)

            if not sheet_names:
                logger.warning(f"No sheets found in Excel file: {excel_path}")
                return []

            # Default to converting all sheets if no prompt callback
            sheets_to_convert = sheet_names

            # If prompt callback is provided, ask which sheets to convert
            if prompt_callback and callable(prompt_callback):
                prompt_message = f"Select sheets to convert from Excel file '{original_filename}':"
                sheets_to_convert = prompt_callback(prompt_message, sheet_names)

                # If no sheets selected, return empty list
                if not sheets_to_convert:
                    logger.info(f"No sheets selected for conversion from {excel_path}")
                    return []

            # Convert selected sheets
            for sheet_name in sheets_to_convert:
                if sheet_name not in sheet_names:
                    logger.warning(f"Sheet '{sheet_name}' not found in {excel_path}")
                    continue

                # Create CSV filename
                base_name = os.path.splitext(secure_filename)[0]
                # Sanitize sheet name for filename
                safe_sheet_name = "".join(c if c.isalnum() else "_" for c in sheet_name)
                csv_filename = f"{base_name}_{safe_sheet_name}.csv"
                csv_path = os.path.join(self.output_dir, csv_filename)

                # Convert sheet to CSV
                self._convert_sheet_to_csv(excel_path, sheet_name, csv_path)

                conversion_results.append(
                    {
                        "sheet_name": sheet_name,
                        "csv_filename": csv_filename,
                        "csv_path": csv_path,
                        "source_filename": secure_filename,
                        "original_excel_filename": original_filename,
                        "email_id": email_id,
                    }
                )

            return conversion_results

        except Exception as e:
            logger.error(f"Failed to convert Excel file {excel_path}: {str(e)}")
            raise ExcelConversionError(str(e), excel_path)

    def _get_sheet_names(self, excel_path: str) -> List[str]:
        """
        Get the names of worksheets in an Excel workbook.
        
        Args:
            excel_path: Path to the Excel file
            
        Returns:
            List of sheet names
            
        Raises:
            ExcelConversionError: If sheet names cannot be retrieved
        """
        try:
            # Try openpyxl first (.xlsx files)
            try:
                workbook = load_workbook(excel_path, read_only=True)
                return [str(name) for name in workbook.sheetnames]  # Ensure all values are strings
            except Exception as e:
                # Fall back to pandas for older Excel formats (.xls)
                xl = pd.ExcelFile(excel_path)
                return [str(name) for name in xl.sheet_names]  # Ensure all values are strings
        except Exception as e:
            logger.error(f"Failed to get sheet names from {excel_path}: {str(e)}")
            raise ExcelConversionError(f"Failed to get sheet names: {str(e)}", excel_path)

    def _convert_sheet_to_csv(self, excel_path: str, sheet_name: str, csv_path: str) -> None:
        """
        Convert a single Excel worksheet to CSV.

        Args:
            excel_path: Path to the Excel file
            sheet_name: Name of the worksheet to convert
            csv_path: Path to save the CSV file

        Raises:
            ExcelConversionError: If conversion fails
        """
        try:
            # Read Excel sheet
            df = pd.read_excel(excel_path, sheet_name=sheet_name)

            # Write to CSV
            df.to_csv(csv_path, index=False, encoding="utf-8")

            logger.info(f"Converted sheet '{sheet_name}' to CSV: {csv_path}")

        except Exception as e:
            logger.error(f"Failed to convert sheet '{sheet_name}' to CSV: {str(e)}")
            raise ExcelConversionError(str(e), excel_path, sheet_name)

    def detect_excel_file(self, content: bytes) -> bool:
        """
        Detect if content is an Excel file based on file signature.

        Args:
            content: File content as bytes

        Returns:
            True if content appears to be an Excel file
        """
        # Excel file signatures
        xlsx_sig = b"PK\x03\x04"  # XLSX files are ZIP archives
        xls_sig = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"  # XLS files signature

        if content.startswith(xlsx_sig) or content.startswith(xls_sig):
            return True

        return False
    
    def convert_standalone(self, file_path: Path, output_dir: Path, 
                          options: Optional[Dict[str, Any]] = None) -> Path:
        """
        Convert an Excel file standalone without email context.
        
        Args:
            file_path: Path to the Excel file to convert
            output_dir: Directory where output should be saved
            options: Optional conversion options
            
        Returns:
            Path to the main output file
            
        Raises:
            ExcelConversionError: If conversion fails
        """
        from pathlib import Path as PathlibPath
        
        # Ensure output directory exists
        output_dir = PathlibPath(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
        
        # Convert to the specified output directory
        original_output_dir = self.output_dir
        self.output_dir = str(output_dir)
        
        try:
            # Generate unique identifiers for standalone conversion
            email_id = f"standalone_{file_path.stem}"
            secure_filename = file_path.name
            
            # Convert all sheets
            results = self.convert_excel_to_csv(
                excel_path=str(file_path),
                original_filename=file_path.name,
                secure_filename=secure_filename,
                email_id=email_id,
                prompt_callback=None  # Convert all sheets by default
            )
            
            if not results:
                raise ExcelConversionError("No sheets were converted", str(file_path))
            
            # Return path to the first converted file
            return PathlibPath(results[0]['csv_path'])
            
        finally:
            # Restore original output directory
            self.output_dir = original_output_dir
